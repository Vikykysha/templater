import os, pprint
import openpyxl
import sys
import datetime
import glob
import re
import shutil
import time
import distutils
from distutils import dir_util

def row_number_search(column_number, row_name_of_col, sheet_object, search_name_of_row):
	while True:
		if str(sheet_object.cell(column = column_number, row = row_name_of_col).value) == search_name_of_row:
			return row_name_of_col
		row_name_of_col = row_name_of_col + 1
		if row_name_of_col > 30:
			row_name_of_col = 'MISSED'
			return row_name_of_col


script_dir = os.path.dirname(__file__)
map_files = os.path.join(script_dir, 'mapping')
#templates_dir = os.path.join(script_dir, 'templates')
result_dir = os.path.join(script_dir,'results')
empty_files = os.path.join(script_dir,'templater_empty')

excel_f = glob.glob(os.path.join(map_files,'*.xlsx'))[0] # glob.glob возвращает лист файлов, подходящих под паттерн, нам надо будет взять только один из них, берем под индексом 0
empty_f = glob.glob(os.path.join(empty_files,'*.xlsx'))[0]
#print(excel_f,empty_f) #debug
table_name = input("Введите название таблицы, которую хотите протестировать\n").lower()
test_name = ''
test_name = input("Введите название теста, если хотите назвать тест другим именем, отличным от имени таблицы.Если тест будет называться отлично от имени таблицы, просто нажмите Enter\n")

where_restrict = 'and ' + input("Напишите ограничения по where, если таковые имеются, where писать не надо\n")

sandbox = ''
sandbox  = input("Напишите номер схемы, если есть, иначе просто нажмите Enter\n")

wb = openpyxl.load_workbook(excel_f, data_only = False, guess_types = False)
map_pattern = re.compile(r'\w*[M,m]apping\w*')
sheet_list = wb.sheetnames
map_sheet  =wb[str([m.group(0) for l in sheet_list for m in [map_pattern.search(l)] if m][0])] #получаем лист с маппингом из excel
#print(map_sheet)

exclude_list = ['PPN_DT','PPN_TM','SRC_STM_ID','PCS_TASK_ID','OPRN_TP','EFF_DT','END_DT']

attr_col_name_list = ['code','физический атрибут'] # список возможных наименований столбца с названиями атрибутов
tabl_col_name_list = ['table','Физическая сущность']
shema_col_name_list = ['schema']
data_type_name_list = ['data type']
length_name_list = ['length']
key_name_list = ['primary key', 'pk']
not_null_name_list = ['not_null'] # пока не используем
required_name_list = ['required']
version_name_list = ['version','версия']

max_row = map_sheet.max_row
#print('max_row = ', max_row) #debug
#ищем номер маппинга
file_name = os.path.basename(excel_f)

map_version = file_name[file_name.rfind('_', 0, file_name.rfind('_')) + 1 : file_name.rfind('_')]

#ПОИСК СТОЛБОЦВ####################################################################
#ищем номер столбца с названием атрибутов
def find_column_in_map(column_number, row, sheet, list_search, raised_exc = True):
	while True:
		#print(sheet.cell(column = column_number, row = row).value)
		#print(list_search)

		if str(sheet.cell(column = column_number, row = row).value).lower() in list_search:
			#print(column_number)
			return column_number
		column_number = column_number + 1
		if column_number > 30:
			if raised_exc:
				raise Exception('This list doesn\'t have the necessary column to build the tests')	
			else:	
				column_number = 'MISSED'
				return column_number

column_version = 1
column_version = find_column_in_map(column_version, 1, map_sheet, version_name_list, False)
#ищем номер столбца со списком атрибутов
column_attr = 1
column_attr = find_column_in_map(column_attr, 1, map_sheet, attr_col_name_list, True)		
#ищем номер столбца с названием таблицы
column_table = 1
column_table = find_column_in_map(column_table, 1, map_sheet, tabl_col_name_list, True)				
#ищем номер столбца с названием схемы
column_schema = 1
column_schema = find_column_in_map(column_schema, 1, map_sheet, shema_col_name_list, True)	
#ищем номер столбца с названием типов данных
column_data_type = 1
column_data_type = find_column_in_map(column_data_type, 1, map_sheet, data_type_name_list, False)		
#ищем номер столбца с названием длины
column_length = 1
column_length = find_column_in_map(column_length, 1, map_sheet, length_name_list, False)			
#ищем номер столбца с названием ключей
column_key = 1
column_key = find_column_in_map(column_key, 1, map_sheet, key_name_list, True)		
#ищем номер столбца с названием not_null
column_not_null = 1
column_not_null = find_column_in_map(column_not_null, 1, map_sheet, not_null_name_list, False)	
#ищем номер столбца с названием обязательным заполнением
column_required = 1
column_required = find_column_in_map(column_required, 1, map_sheet, required_name_list, False)

		
#####################################################################################
#debug
#print(column_version, column_attr, column_table, column_schema, column_data_type, column_length, column_key, column_not_null, column_required)

property_dict = {'key' : None, 'not_null' : None, 'exclude' : None, 'required' : None, 'object_id' : None, 'length': None,'column_data_type':None, 'version':None}

#находим все атрибуты интересующей таблицы
attr_list = []
for row in range(2, max_row):
	#print(column_table,row,map_sheet.cell(column = column_table, row = row).value,table_name.strip())
	if map_sheet.cell(column = column_table, row = row).value.strip().lower() == table_name.strip():
		attr_list.append(map_sheet.cell(column = column_attr, row = row).value.strip())
#print(attr_list)#debug	
#составляем список атрибутов и их характеристик
#dict_list = [property_dict] * len(attr_list) - неправильно, тогда делаем просто несколько копий одного и того же объекта, при изменении одного поменяются все
dict_list = []
for _ in range(len(attr_list)):
	dict_list.append({'key' : None, 'not_null' : None, 'exclude' : None, 'required' : None, 'object_id' : None, 'length': None,'column_data_type':None, 'version':None})

attr_prop_dict = dict(zip(attr_list,dict_list))
#print(attr_prop_dict)#debug


#находим название схемы
schema_name = None
for row in range(2, max_row):
	if map_sheet.cell(column = column_table, row = row).value.strip().lower() == table_name.strip():
		schema_name = map_sheet.cell(column = column_schema, row = row).value.strip()
		break
#print(schema_name)

###############ЗАПОЛНЕНИЕ ХАРАКТЕРИСТИК АТРИБУТОВ#############################################################
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!	
#заполнение характеристик атрибутов

for row in range(2, max_row):
	if attr_prop_dict.get(map_sheet.cell(column = column_attr, row = row).value) != None and str(map_sheet.cell(column = column_table, row = row).value).lower().strip() == table_name:
		attr_prop_dict[map_sheet.cell(column = column_attr, row = row).value]['key'] = map_sheet.cell(column = column_key, row = row).value
		if column_required != 'MISSED':
			attr_prop_dict[map_sheet.cell(column = column_attr, row = row).value]['required'] = map_sheet.cell(column = column_required, row = row).value
		if column_length != 'MISSED':
			attr_prop_dict[map_sheet.cell(column = column_attr, row = row).value]['length'] = map_sheet.cell(column = column_length, row = row).value
		if column_data_type != 'MISSED':
			attr_prop_dict[map_sheet.cell(column = column_attr, row = row).value]['column_data_type'] = map_sheet.cell(column = column_data_type, row = row).value
		if map_sheet.cell(column = column_attr, row = row).value.strip() in exclude_list:
			attr_prop_dict[map_sheet.cell(column = column_attr, row = row).value]['exclude'] = 'yes'
		if map_sheet.cell(column = column_key, row = row).value is not None: #and map_sheet.cell(column = column_not_null, row = row).value != '':
			attr_prop_dict[map_sheet.cell(column = column_attr, row = row).value]['not_null'] = 'yes' 
		if column_version != 'MISSED':
			attr_prop_dict[map_sheet.cell(column = column_attr, row = row).value]['version'] = map_sheet.cell(column = column_version, row = row).value 


		
		



wb = openpyxl.load_workbook(os.path.join(empty_files,'Templater_empty.xlsx'),data_only=False, guess_types=False)
sheet_fileds = wb['FIELDS']
attr_qunt = len(attr_list)
column = 2
cnt = 0

for row in range(3,attr_qunt + 3):
	cell = sheet_fileds.cell(row = row, column = column)
	cell.value = attr_list[cnt]
	cnt +=1
	

def find_column_in_map_field_lst(column_number, row, sheet, raised_exc = True, value_to = None):
	while True:
		if str(sheet.cell(column = column_number, row = row).value).lower() == value_to.lower():
			return column_number
		column_number = column_number + 1
		if column_number > 30:
			if raised_exc:
				raise Exception('This list doesn\'t have the necessary column')	
			else:	
				column_number = 'MISSED'
				return column_number

#Ищем номера столбцов для заполнения
column_excl = 1
column_excl = find_column_in_map_field_lst(column_excl, 2, sheet_fileds, False, 'exclude' )	
column_key = 1
column_key = find_column_in_map_field_lst(column_key, 2, sheet_fileds, False, 'key' )	
column_not_null = 1
column_not_null = find_column_in_map_field_lst(column_not_null, 2, sheet_fileds, False, 'not_null' )
column_object_id = 1
column_object_id  = find_column_in_map_field_lst(column_object_id, 2, sheet_fileds,  False, 'object_id' )		
column_not_all_null_attr = 1
column_not_all_null_attr = find_column_in_map_field_lst(column_not_all_null_attr, 2, sheet_fileds,  False, 'NOT_ALL_NULL_ATR' )

#Заполняем пустой темлейтер значениями
column = 2
print(column_not_all_null_attr)
for row in range(3,attr_qunt + 3):
	#print('excl val ' , attr_prop_dict[sheet_fileds.cell(column = column, row = row).value])
	if column_excl != 'MISSED'  and attr_prop_dict[sheet_fileds.cell(column = column, row = row).value]['exclude'] == 'yes':
		sheet_fileds.cell(column = column_excl, row = row).value = 'X'
	if column_key != 'MISSED':
		sheet_fileds.cell(column = column_key, row = row).value = attr_prop_dict[sheet_fileds.cell(column = column, row = row).value]['key']
	if column_not_null != 'MISSED' and attr_prop_dict[sheet_fileds.cell(column = column, row = row).value]['not_null'] == 'yes':
		sheet_fileds.cell(column = column_not_null, row = row).value = 'X'
	if column_object_id != 'MISSED':
		sheet_fileds.cell(column = column_object_id, row = row).value = attr_prop_dict[sheet_fileds.cell(column = column, row = row).value]['object_id']
	if column_not_all_null_attr != 'MISSED':
		sheet_fileds.cell(column = column_not_all_null_attr, row = row).value = attr_prop_dict[sheet_fileds.cell(column = column, row = row).value]['required']


sheet_gener = wb['GENER']

source_schema_row = 1
source_table_row = 1
source_where_row = 1
target_schema_row = 1
target_where_row = 1
target_table_row = 1
etalon_schema_row = 1
etalon_table_row = 1
etalon_where_row = 1
log_schema_row = 1
log_table_row = 1
loading_dt_row = 1
release_row = 1
map_version_row = 1
test_name_row = 1
sandbox_row = 1


source_schema_row = row_number_search(2, source_schema_row, sheet_gener, 'SOURCE_SCHEMA')
source_table_row = row_number_search(2, source_table_row, sheet_gener, 'SOURCE_TABLE')
source_where_row = row_number_search(2, source_where_row, sheet_gener, 'SOURCE_WHERE')
target_schema_row = row_number_search(2, target_schema_row, sheet_gener, 'TARGET_SCHEMA')
target_table_row = row_number_search(2, target_table_row, sheet_gener, 'TARGET_TABLE')
target_where_row = row_number_search(2, target_where_row, sheet_gener, 'TARGET_WHERE')
etalon_schema_row = row_number_search(2, etalon_schema_row, sheet_gener, 'ETALON_SCHEMA')
etalon_table_row = row_number_search(2, etalon_table_row, sheet_gener, 'ETALON_TABLE')
etalon_where_row = row_number_search(2, etalon_where_row, sheet_gener, 'ETALON_WHERE')
log_schema_row = row_number_search(2, log_schema_row, sheet_gener, 'LOG_SCHEMA')
log_table_row = row_number_search(2, log_table_row, sheet_gener, 'LOG_TABLE')
loading_dt_row = row_number_search(2, loading_dt_row, sheet_gener, 'LOADING_DT')
release_row = row_number_search(2, release_row, sheet_gener, 'RELEASE')
map_version_row = row_number_search(2, map_version_row, sheet_gener, 'MAP_VERSION')
test_name_row = row_number_search(2, test_name_row, sheet_gener, 'TEST_NAME')
sandbox_row = row_number_search(2, sandbox_row, sheet_gener, 'SANDBOX')

#print('table_name' + table_name, target_table_row)

if target_table_row != 'MISSED':
	sheet_gener.cell(column = 3, row = target_table_row).value = table_name
	sheet_gener.cell(column = 3, row = etalon_table_row).value = table_name
else:
	sheet_gener.cell(column = 3, row = target_table_row).value = ''
	sheet_gener.cell(column = 3, row = etalon_table_row).value = ''
if loading_dt_row != 'MISSED':
	sheet_gener.cell(column = 3, row = loading_dt_row).value = '#LOADING_DT#'
else:
	sheet_gener.cell(column = 3, row = loading_dt_row).value = ''
if log_schema_row != 'MISSED':
	sheet_gener.cell(column = 3, row = log_schema_row).value = 'T_TEST_LOG'
else:
	sheet_gener.cell(column = 3, row = log_schema_row).value = ''
if test_name != '':
	sheet_gener.cell(column = 3, row = test_name_row).value = test_name
else:
	sheet_gener.cell(column = 3, row = test_name_row).value = table_name

if log_schema_row != 'MISSED':
	sheet_gener.cell(column = 3, row = map_version_row).value = map_version
else:
	sheet_gener.cell(column = 3, row = map_version_row).value = ''
if target_where_row != 'MISSED' and where_restrict.strip() != 'and':
	sheet_gener.cell(column = 3, row = target_where_row).value = where_restrict
else:
	sheet_gener.cell(column = 3, row = target_where_row).value = ''
if etalon_where_row != 'MISSED' and where_restrict.strip() != 'and':
	sheet_gener.cell(column = 3, row = etalon_where_row).value = where_restrict
else:
	sheet_gener.cell(column = 3, row = etalon_where_row).value = ''
if sandbox != '':
	sheet_gener.cell(column = 3, row = sandbox_row).value = schema_name +'_' + sandbox
	sheet_gener.cell(column = 3, row = sandbox_row).value = sandbox
	schema_name = schema_name + '_' + sandbox
else:
	sheet_gener.cell(column = 3, row = sandbox_row).value = schema_name
	sheet_gener.cell(column = 3, row = sandbox_row).value = ''

release_list = []
for key in attr_prop_dict:
	release_list.append(int(attr_prop_dict[key]['version']))
#print(release_list) #debug
release = max(release_list)

if release_row != 'MISSED':
	sheet_gener.cell(column = 3, row = release_row).value = release



#ПРОВОДИМ ЗАПОЛНЕНИЕ ТЕСТА НА НУЛЕВЫЕ АТРИБУТЫ
null_test = 'select '
null_attr_list = []
for key, value in attr_prop_dict.items():
	if attr_prop_dict[key]['required'] != None :
		null_attr_list.append('case when ' + key + ' = 0 then \'FAIL\' else \'PASSED\' end as ' + key + ' \n')
sel_null_attr_list = []
for key, value in attr_prop_dict.items():
	if attr_prop_dict[key]['required'] != None :
		sel_null_attr_list.append('( select count( ' + key + ' ) ' + key + ' from '+schema_name+'.' + table_name +  ' ) ' + table_name + '\n')
null_test += ','.join(null_attr_list) + ' from \n' + ','.join(sel_null_attr_list) + ';\n'
visual_test = wb.create_sheet(title = 'VISUAL_TESTS') #18.10.2018
cell1 = visual_test.cell(row = 1, column = 1)
cell1.value = null_test 


cnt_del_r = 3
while True: #добавление функционала от 15.10.2018
	if  not sheet_fileds['B' + str(cnt_del_r)].value:
		break
	else:
		cnt_del_r += 1


print(cnt_del_r)
sheet_fileds.delete_rows(cnt_del_r,125)
new_File_name = table_name + '_' + str(release)
print("Сохраняем файл ..")
os.chdir(result_dir)
wb.save(new_File_name + '.xlsx')



print("Готово")




